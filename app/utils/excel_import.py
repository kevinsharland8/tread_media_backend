import openpyxl
from openpyxl_image_loader import SheetImageLoader
import os
from utils.upload_google import upload_to_bucket
from db import get_postgres, insert_with_error_handling, fetch_with_error_handling
from pydantic import ValidationError
from models import (
    Event,
    EventType,
    Provice,
    JunctionTable,
    EventImage_id,
    EventDistance_id,
    ErrorCreate_id,
)
from datetime import datetime


async def insert_data_event_table(event: Event, event_image: str) -> Event:
    if not event:
        raise ValueError("No fields provided for insert events")
    db_pool = await get_postgres()
    query = """
    INSERT INTO events (name, description, start_date, end_date, city, province_id, event_website, organizer, active, map_link, multi_day)
    VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11)
    RETURNING id, name, description, start_date, end_date, city, province_id, event_website, organizer, active, map_link, multi_day;
    """
    query_filters = [
        event.name,
        event.description,
        event.start_date,
        event.end_date,
        event.city,
        event.province_id,
        event.event_website,
        event.organizer,
        event.active,
        event.map_link,
        event.multi_day,
    ]
    unique_event_id = await insert_with_error_handling(
        db_pool=db_pool, query=query, query_filters=query_filters, model=Event
    )

    return unique_event_id["result"]


async def insert_data_images_table(event_id: int, event_image: str):
    if not event_id:
        raise ValueError("No fields provided for insert events")
    db_pool = await get_postgres()
    query = """
    INSERT INTO event_images (event_id, headline, url) values ($1, $2, $3)
    RETURNING id, event_id, headline, url;
    """

    query_filters = [
        event_id,
        True,
        f"https://storage.googleapis.com/tread_media_images/{event_image}",
    ]

    unique_event_image_id = await insert_with_error_handling(
        db_pool=db_pool, query=query, query_filters=query_filters, model=EventImage_id
    )
    return {"results": [event_id, unique_event_image_id]}


async def insert_event_junction(event_id: int, event_type: int):
    if event_id is None or event_type is None:
        raise ValueError("No fields provided for insert events junction")
    db_pool = await get_postgres()
    query = """
    INSERT INTO event_event_types_junction (event_id, event_type_id)
    VALUES ($1, $2)
    RETURNING id, event_id, event_type_id;
    """
    query_filters = [event_id, event_type]
    await insert_with_error_handling(
        db_pool=db_pool, query=query, query_filters=query_filters, model=JunctionTable
    )

    return "unique_event_id"


async def insert_event_distances(event_id: int, day: int, distance: int):
    if event_id is None or day is None or distance is None:
        raise ValueError("No fields provided for insert dinstances")
    db_pool = await get_postgres()
    query = """
    INSERT INTO event_distances (event_id, day, distance)
    VALUES ($1, $2, $3)
    RETURNING id, event_id, day, distance;
    """
    query_filters = [event_id, day, distance]
    await insert_with_error_handling(
        db_pool=db_pool,
        query=query,
        query_filters=query_filters,
        model=EventDistance_id,
    )

    return "unique_event_id"


async def insert_error(row_number, error, inserted, run_date):
    if row_number is None or error is None or inserted is None:
        raise ValueError("some or no fields were provided for inserting")
    db_pool = await get_postgres()
    query = """
    INSERT INTO errors (row_number, error, inserted, run_date)
    VALUES ($1, $2, $3, $4)
    RETURNING id, row_number, error, inserted, run_date;
    """
    query_filters = [row_number, error, inserted, run_date]
    await insert_with_error_handling(
        db_pool=db_pool,
        query=query,
        query_filters=query_filters,
        model=ErrorCreate_id,
    )
    return "error inserted"


async def get_event_type(name):
    db_pool = await get_postgres()
    query = """
    select id, type from event_types where type = $1;
    """
    data_return = await fetch_with_error_handling(
        db_pool=db_pool, query=query, query_filters=name.strip(), model=EventType
    )
    return data_return.id


async def get_province(province):
    db_pool = await get_postgres()
    query = """
    select id, p_name from provinces where p_name = $1;
    """
    data_return = await fetch_with_error_handling(
        db_pool=db_pool, query=query, query_filters=province.strip(), model=Provice
    )
    return data_return.id


async def extract_from_file(file):
    workbook = openpyxl.load_workbook(file)
    # workbook = openpyxl.load_workbook("/home/kevin/projects/tread-events-python/app/utils/excel_with_image.xlsx")
    sheet = workbook["testing"]
    output_directory = "/home/kevin/projects/tread-events-python/images/"
    # Initialize image loader
    image_loader = SheetImageLoader(sheet)
    progress = True

    try:
        # get the header column
        headers = [cell.value for cell in sheet[1]]
        # min_row = the starting row, max_row = last row with data, min_col = starting column, max_col = last column with data
        for row in sheet.iter_rows(
            min_row=2, max_row=sheet.max_row, min_col=1, max_col=14
        ):
            # used for the image
            row_num = row[0].row  # Get the row number
            row_values = [cell.value for cell in row]
            row_dict = dict(zip(headers, row_values))
            print(
                "----------------------------------------------------------------------------------------------------------------------------"
            )
            print(row_num, row_values)
            # convert the string of event types into a list
            try:
                eti = row_dict["event_type_id"]
                if "," in eti:
                    update_event_type = eti.split(",")
                else:
                    update_event_type = [eti]
                row_dict.update({"event_type_id": update_event_type})
            except Exception as e:
                error_message = (
                    "event type was not provided, please check all data is povided"
                )
                print(error_message)
                await insert_error(row_num, error_message, False, datetime.now())
                continue

            # convert the string of distances into a list
            if row_dict["distances"] is not None:
                ud = str(row_dict["distances"])
                if "," in ud:
                    update_distance = ud.split(",")
                else:
                    update_distance = [ud]
                row_dict.update({"distances": update_distance})
            else:
                error_message = (
                    "distance was not provided, please check all data is povided"
                )
                print(error_message)
                await insert_error(row_num, error_message, False, datetime.now())
                continue

            # get the id of the province from the province table
            try:
                provice_id = await get_province(row_dict["province"])
                row_dict.update({"province": provice_id})
            except Exception as e:
                row_dict.update({"province": None})

            # data for the events table
            insert_data_event = {
                "name": row_dict["name"],
                "description": row_dict["description"],
                "start_date": row_dict["start_date"],
                "end_date": row_dict["end_date"],
                "city": row_dict["city"],
                "province_id": row_dict["province"],
                "event_website": row_dict["event_website"],
                "organizer": row_dict["organizer"],
                "active": row_dict["active"],
                "map_link": row_dict["map_link"],
                "multi_day": row_dict["multi_day"],
            }
            try:
                event = Event(**insert_data_event)
            except ValidationError as e:
                error_message = f"{e.errors()[0]['loc'][0]} was not provided, please check all data is povided"
                print(error_message)
                await insert_error(row_num, error_message, False, datetime.now())
                continue
            # need to update this if the image column changes
            image_cell = f"N{row_num}"
            if image_loader.image_in(image_cell):
                image = image_loader.get(image_cell)
                clean_date = row_dict["start_date"].date()
                clean_name = f"{row_dict['name']}_{clean_date}.png"
                clean_name = "".join(clean_name.split())
                complete_path_image = os.path.join(output_directory, clean_name)
                image.save(complete_path_image)
                try:
                    upload_to_bucket(complete_path_image, clean_name)
                    use_default_image = False
                except Exception as e:
                    print(f"there is a n error copying it to the bucket {e}")
            else:
                error_message = "image was not provided, data was still inserted"
                print(error_message)
                await insert_error(row_num, error_message, True, datetime.now())
                use_default_image = True
                # continue

            # insert into the events table
            try:
                insert_into_tables = await insert_data_event_table(event, clean_name)
            except Exception as e:
                progress = False
                error_message = f"{e}"
                print(error_message)
                await insert_error(row_num, error_message, False, datetime.now())
                continue

            # insert into junction table
            # get the id of the event from the event_types table
            try:
                for event_type_ in row_dict["event_type_id"]:
                    insert_event_type_id = await get_event_type(event_type_)
                    await insert_event_junction(
                        insert_into_tables, insert_event_type_id
                    )
            except Exception as e:
                progress = False
                print(
                    f"there is a n error inserting the data into events junction table {e}"
                )

            # insert into the images table
            if use_default_image:
                clean_name = "default.png"
            try:
                await insert_data_images_table(insert_into_tables, clean_name)
            except Exception as e:
                progress = False
                print(f"there is a n error inserting the data into images {e}")

            # insert the distance
            # if multi is true the distances will be the distance per day
            # if it is false it is different distances on the day
            try:
                multi_true_false = row_dict["multi_day"]
                counter = 0
                if not multi_true_false:
                    for event_distance_ in row_dict["distances"]:
                        await insert_event_distances(
                            int(float(insert_into_tables)),
                            1,
                            int(float(event_distance_)),
                        )
                else:
                    for event_distance_ in row_dict["distances"]:
                        counter = counter + 1
                        await insert_event_distances(
                            int(insert_into_tables),
                            int(float(counter)),
                            int(float(event_distance_)),
                        )
            except Exception as e:
                progress = False
                print(f"there is a n error inserting the data into distance table {e}")

            # progress = False
        return progress

    except Exception as e:
        print(f"error Processing file, {e}")
